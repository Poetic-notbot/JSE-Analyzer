"""Parser for MoA weekly commodity-price workbooks (.xlsx).

These files are NOT a uniform schema — each report type has a different layout
(verified against real files). We detect the report type, then branch:

* **Farmgate / Rural Retail** — parish-based. Multi-row header where parish names
  are MERGED cells each spanning 5 metric columns
  (Low, High, Most Frequent, Supply, Grade). Columns A/B are Commodity and
  Variety/Source. Missing values are the string "-". A footer row
  ("Prepared on ...") is dropped. Output is unpivoted to a tidy long form.
* **Retail / Wholesale** — Kingston Metropolitan Region supermarket layout with
  an "Average Price" column plus many per-store location columns. We take the
  Average Price as the representative price.

The parser returns a long DataFrame with columns:
    commodity, variety, location, price_jmd, price_type, week_ending
which ministry_ingestion normalizes into price_records (J$/kg already per kg).
Any layout it does not recognise yields an empty frame so the caller falls back
to the generic tabular parser.
"""

from __future__ import annotations

import io
import re

import pandas as pd

_MONTHS = {m: i for i, m in enumerate(
    ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct",
     "nov", "dec"], start=1)}

METRIC_HEADERS = {"low", "high", "most frequent", "supply", "grade"}

# report keyword (checked longest-first) -> price_type
_REPORT_TYPES = [
    ("rural retail meat", "rural_retail"),
    ("retail meat", "retail"),
    ("rural retail", "rural_retail"),
    ("urban municipal", "urban_retail"),
    ("wholesale", "wholesale"),
    ("farmgate", "farmgate"),
    ("retail", "retail"),
]
_PARISH_LAYOUTS = {"farmgate", "rural_retail"}


def _clean(v) -> str | None:
    if v is None:
        return None
    s = re.sub(r"\s+", " ", str(v).replace("\n", " ")).strip()
    return s or None


def parse_week_ending(text: str | None) -> str | None:
    if not text:
        return None
    m = re.search(r"week ending\s+([A-Za-z]+)\.?\s+(\d{1,2}),?\s+(\d{4})",
                  text, re.IGNORECASE)
    if m:
        mo = _MONTHS.get(m.group(1).lower()[:3])
        if mo:
            return f"{int(m.group(3)):04d}-{mo:02d}-{int(m.group(2)):02d}"
    # numeric MM.DD.YYYY (e.g. from filename)
    m = re.search(r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})", text)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    return None


def detect_report_type(filename: str, title_text: str) -> str | None:
    hay = f"{filename or ''} {title_text or ''}".lower()
    for kw, ptype in _REPORT_TYPES:
        if kw in hay:
            return ptype
    return None


def _to_price(v) -> float | None:
    s = _clean(v)
    if s is None or s in ("-", "--", "n/a", "na"):
        return None
    s = re.sub(r"[^0-9.]", "", s)
    try:
        return float(s) if s else None
    except ValueError:
        return None


def parse_workbook(content: bytes, filename: str) -> pd.DataFrame:
    """Return a tidy long DataFrame of weekly prices, or empty if unrecognised."""
    try:
        import openpyxl  # lazy: heavy import
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)
    except Exception:  # noqa: BLE001 - not an xlsx / unreadable -> caller falls back
        return pd.DataFrame()

    ws = wb.worksheets[0]
    for cand in wb.worksheets:
        if "report" in (cand.title or "").lower():
            ws = cand
            break

    # Gather the first several rows as cleaned text for detection.
    grid: list[list] = []
    for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        grid.append(list(r))
    if not grid:
        return pd.DataFrame()

    top_text = " ".join(_clean(c) or "" for row in grid[:4] for c in row)
    week_ending = parse_week_ending(top_text) or parse_week_ending(filename)
    report_type = detect_report_type(filename, top_text)
    if report_type is None:
        return pd.DataFrame()

    if report_type in _PARISH_LAYOUTS:
        return _parse_parish_layout(ws, grid, report_type, week_ending)
    return _parse_supermarket_layout(grid, report_type, week_ending)


# --------------------------------------------------------------------------- #
# Parish (Farmgate / Rural Retail) layout with merged parish header cells.
# --------------------------------------------------------------------------- #
def _find_subheader_row(grid: list[list]) -> int | None:
    for i, row in enumerate(grid):
        cells = {(_clean(c) or "").lower() for c in row}
        if len({"low", "high", "most frequent"} & cells) >= 2:
            return i
    return None


def _parse_parish_layout(ws, grid, report_type, week_ending) -> pd.DataFrame:
    sub_i = _find_subheader_row(grid)
    if sub_i is None or sub_i < 1:
        return pd.DataFrame()
    parish_i = sub_i - 1
    sub_row = [(_clean(c) or "").lower() for c in grid[sub_i]]

    # Map each column -> parish, expanding merged ranges on the parish row.
    col_parish: dict[int, str] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_row == parish_i + 1:  # openpyxl is 1-indexed
            name = _clean(ws.cell(rng.min_row, rng.min_col).value)
            if name:
                for col in range(rng.min_col, rng.max_col + 1):
                    col_parish[col - 1] = name
    # also any non-merged parish cells
    for c, val in enumerate(grid[parish_i]):
        name = _clean(val)
        if name and c not in col_parish and c >= 2:
            col_parish[c] = name

    # "Most Frequent" column per parish block.
    mf_cols = {c: p for c, p in col_parish.items()
               if c < len(sub_row) and sub_row[c] == "most frequent"}
    if not mf_cols:
        return pd.DataFrame()

    rows = []
    for r in range(sub_i + 1, len(grid)):
        commodity = _clean(grid[r][0]) if grid[r] else None
        if not commodity or commodity.lower().startswith("prepared on"):
            continue
        variety = _clean(grid[r][1]) if len(grid[r]) > 1 else None
        for c, parish in mf_cols.items():
            price = _to_price(grid[r][c]) if c < len(grid[r]) else None
            if price is None or price <= 0:
                continue
            rows.append({"commodity": commodity, "variety": variety,
                         "location": parish, "price_jmd": price,
                         "price_type": report_type, "week_ending": week_ending})
    return pd.DataFrame(rows)


# --------------------------------------------------------------------------- #
# Supermarket (Retail / Wholesale) layout — use the Average Price column.
# --------------------------------------------------------------------------- #
def _parse_supermarket_layout(grid, report_type, week_ending) -> pd.DataFrame:
    header_i = None
    avg_c = commodity_c = variety_c = None
    for i, row in enumerate(grid):
        cells = [(_clean(c) or "").lower() for c in row]
        if any("average" in c for c in cells) and any("commodity" in c for c in cells):
            header_i = i
            for c, txt in enumerate(cells):
                if "commodity" in txt:
                    commodity_c = c
                elif "variety" in txt or "source" in txt:
                    variety_c = c
                elif "average" in txt:
                    avg_c = c
            break
    if header_i is None or avg_c is None or commodity_c is None:
        return pd.DataFrame()

    rows = []
    for r in range(header_i + 1, len(grid)):
        row = grid[r]
        commodity = _clean(row[commodity_c]) if commodity_c < len(row) else None
        if not commodity or commodity.lower().startswith("prepared on"):
            continue
        price = _to_price(row[avg_c]) if avg_c < len(row) else None
        if price is None or price <= 0:
            continue
        variety = (_clean(row[variety_c])
                   if variety_c is not None and variety_c < len(row) else None)
        rows.append({"commodity": commodity, "variety": variety,
                     "location": "Kingston Metropolitan Region",
                     "price_jmd": price, "price_type": report_type,
                     "week_ending": week_ending})
    return pd.DataFrame(rows)
