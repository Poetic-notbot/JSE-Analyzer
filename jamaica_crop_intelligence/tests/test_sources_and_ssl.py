"""Tests for the SSL/error-classification fix, FAOSTAT adapter, MoA weekly
XLSX parsers, and the official-supersedes-illustrative rule."""

import io
import json

import pandas as pd
import pytest

from jamaica_crop_intelligence.ingestion import faostat, moa_weekly, official_fetch
from jamaica_crop_intelligence.services.import_service import ImportService
from jamaica_crop_intelligence.services.price_service import PriceService
from jamaica_crop_intelligence.services.production_service import ProductionService


# --------------------------------------------------------------- SSL / errors
def test_ca_bundle_resolves():
    bundle = official_fetch.ca_bundle()
    assert bundle and isinstance(bundle, str)


def test_fetch_never_raises_and_classifies():
    # An unreachable/blocked host must return ok=False with an error_class,
    # never raise.
    res = official_fetch.fetch_url("https://www.moa.gov.jm/does-not-exist.pdf",
                                   timeout=15)
    assert res.ok is False
    assert res.error_class in ("cert_trust", "ssl", "timeout", "proxy_block",
                               "network", "forbidden", "not_found", "http", "other")
    assert res.error


def test_combined_ca_bundle_merges_and_caches():
    import re
    path = official_fetch.combined_ca_bundle()
    assert path and __import__("os").path.exists(path)
    n = len(re.findall(r"BEGIN CERTIFICATE", open(path).read()))
    assert n > 1  # certifi + system/env anchors merged
    assert official_fetch.combined_ca_bundle() == path  # cached


def test_transient_5xx_message_is_retriable():
    for status in (502, 521, 523):
        r = official_fetch._http_error("http://x", status, "down")
        assert r.ok is False
        assert "transient" in r.error.lower() or "temporarily" in r.error.lower()
    # 404 is NOT framed as transient
    r404 = official_fetch._http_error("http://x", 404, "nf")
    assert r404.error_class == "not_found"


def test_error_class_distinguishes_cert_from_network():
    import requests
    cls_cert, msg_cert = official_fetch._classify(
        requests.exceptions.SSLError("certificate verify failed: unable to get "
                                     "local issuer certificate"))
    assert cls_cert == "cert_trust"
    assert "not a network block" in msg_cert.lower()
    cls_to, _ = official_fetch._classify(requests.exceptions.ConnectTimeout())
    assert cls_to == "timeout"
    cls_px, _ = official_fetch._classify(requests.exceptions.ProxyError())
    assert cls_px == "proxy_block"


# ------------------------------------------------------------------- FAOSTAT
FAO_RECORDS = [
    {"Item": "Yams", "Year": 2021, "Value": 180000, "Unit": "t"},
    {"Item": "Yams", "Year": 2022, "Value": 185000, "Unit": "t"},
    {"Item": "Chillies and peppers, green", "Year": 2022, "Value": 7000, "Unit": "t"},
    {"Item": "Some Unknown Crop", "Year": 2022, "Value": 100, "Unit": "t"},
    {"Item": "Tomatoes", "Year": 2022, "Value": -5, "Unit": "t"},  # invalid
]


def test_faostat_url_and_parse():
    url = faostat.build_url("production", [2021, 2022])
    assert "area=109" in url and "element=5510" in url and "2021,2022" in url
    recs = faostat.parse_json(json.dumps({"data": FAO_RECORDS}).encode())
    assert len(recs) == 5


def test_faostat_maps_taxonomy_and_validates(repo):
    prev = faostat.build_preview(FAO_RECORDS, "production", repo)
    assert prev["kind"] == "production"
    crops = set(prev["normalized"]["crop_raw"])
    assert "Yams" in crops and "Chillies and peppers, green" in crops
    # unknown flagged, negative dropped
    assert any(c["raw"] == "Some Unknown Crop" for c in prev["crop_review"])
    assert (prev["normalized"]["tonnes"] > 0).all()
    # committed rows are official + annual (quarter NULL)
    imp = ImportService(repo)
    sfid, _ = imp.register_source_file("faostat.json",
                                       json.dumps(FAO_RECORDS).encode(),
                                       "production", source_name="FAOSTAT")
    res = imp.commit_preview(sfid, "production", prev["normalized"],
                             prev["quality_issues"])
    assert res["rows_committed"] == 3
    row = repo.query_one("SELECT quarter, provenance FROM production_records "
                         "WHERE provenance='official_observed' LIMIT 1")
    assert row["quarter"] is None


# ------------------------------------------------------------- MoA weekly xlsx
def _farmgate_workbook() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Farmgate Report"
    ws["A1"] = "Farmgate Prices (J$/Kg)"
    ws["A2"] = "Week Ending June 27, 2026"
    ws["C3"] = "Manchester"; ws.merge_cells("C3:G3")
    ws["H3"] = "St.Andrew"; ws.merge_cells("H3:L3")
    ws["A4"] = "Commodity"; ws["B4"] = "Variety / Source"
    for base in (3, 8):
        for j, h in enumerate(["Low", "High", "Most\nFrequent", "Supply", "Grade"]):
            ws.cell(4, base + j, h)
    rows = [("Tomato", "Local", 180, 220, 200, "Good", "A", 190, 230, 210, "Fair", "A"),
            ("Cabbage", "Local", 120, 160, "-", "Good", "A", 130, 170, 150, "Good", "B")]
    r = 5
    for row in rows:
        ws.cell(r, 1, row[0]); ws.cell(r, 2, row[1])
        for k, v in enumerate(row[2:]):
            ws.cell(r, 3 + k, v)
        r += 1
    ws.cell(r, 1, "Prepared on 06/25/2026 by AMII Branch")
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def _retail_workbook() -> bytes:
    import openpyxl
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Retail Report"
    ws["A1"] = "Kingston Metropolitan Region"
    ws["A2"] = "Retail Prices (J$/Kg) Week Ending June 13, 2026"
    ws["A3"] = "Commodity"; ws["B3"] = "Variety/Source"; ws["C3"] = "Average Price"
    ws["D3"] = "CROSS ROADS"; ws["E3"] = "HALF-WAY TREE"
    ws["D4"] = "Hi-Lo"; ws["E4"] = "Sampars"
    ws["A5"] = "Tomato"; ws["B5"] = "Local"; ws["C5"] = 350; ws["D5"] = 340; ws["E5"] = 360
    ws["A6"] = "Lettuce"; ws["B6"] = "Local"; ws["C6"] = 420; ws["D6"] = 400
    buf = io.BytesIO(); wb.save(buf)
    return buf.getvalue()


def test_moa_farmgate_merged_cells(repo):
    content = _farmgate_workbook()
    df = moa_weekly.parse_workbook(content, "Farmgate 06.27.2026.xlsx")
    assert not df.empty
    assert set(df["price_type"]) == {"farmgate"}
    assert set(df["location"]) == {"Manchester", "St.Andrew"}
    # Cabbage/Manchester "Most Frequent" was "-" -> dropped; St.Andrew kept
    cab = df[df["commodity"] == "Cabbage"]
    assert set(cab["location"]) == {"St.Andrew"}
    assert (df["week_ending"] == "2026-06-27").all()

    imp = ImportService(repo)
    prev = imp.preview(content, "Farmgate 06.27.2026.xlsx", "prices")
    assert prev["columns_detected"]["source"] == "MoA weekly workbook"
    assert prev["summary"]["rows_normalized"] == 3


def test_moa_retail_average_price(repo):
    content = _retail_workbook()
    df = moa_weekly.parse_workbook(content, "Retail 06.13.2026.xlsx")
    assert not df.empty
    assert set(df["price_type"]) == {"retail"}
    assert (df["location"] == "Kingston Metropolitan Region").all()
    tom = df[df["commodity"] == "Tomato"].iloc[0]
    assert tom["price_jmd"] == 350  # average price column


def test_moa_week_ending_and_report_type():
    assert moa_weekly.parse_week_ending("Week Ending June 27, 2026") == "2026-06-27"
    assert moa_weekly.parse_week_ending("Farmgate 06.27.2026.xlsx") == "2026-06-27"
    assert moa_weekly.detect_report_type("Wholesale 01.02.2026.xlsx", "") == "wholesale"
    assert moa_weekly.detect_report_type("x.xlsx", "Rural Retail Prices") == "rural_retail"


# ----------------------------------------------- official supersedes illustrative
def test_official_supersedes_illustrative(repo):
    prod = ProductionService(repo)
    price = PriceService(repo)
    yam = repo.crop_id_by_name("Yellow Yam")
    assert set(prod.production(yam)["provenance"]) == {"illustrative"}

    imp = ImportService(repo)
    recs = [{"Item": "Yams", "Year": 2022, "Value": 185000, "Unit": "t"}]
    prev = faostat.build_preview(recs, "production", repo)
    sfid, _ = imp.register_source_file("f.json", json.dumps(recs).encode(),
                                       "production", source_name="FAOSTAT")
    imp.commit_preview(sfid, "production", prev["normalized"], prev["quality_issues"])

    after = prod.production(yam)
    assert set(after["provenance"]) == {"official_observed"}      # seeds suppressed
    # a crop without official data keeps its illustrative series
    assert "illustrative" in set(prod.production(repo.crop_id_by_name("Tomato"))["provenance"])
    # total_production no longer double counts Yellow Yam
    tp = prod.total_production(2022)
    yam_rows = tp[tp["crop"] == "Yellow Yam"]
    assert len(yam_rows) == 1
